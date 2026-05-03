"""Report export helpers — Excel (openpyxl) and PDF (reportlab)."""

from __future__ import annotations

import io
from datetime import date
from decimal import Decimal
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


# ─── Shared formatting ────────────────────────────────────────────────────────

def _fmt_decimal(v: Decimal | None) -> str:
    if v is None:
        return "-"
    return f"{v:,.0f}"


def _rp(v: Decimal | None) -> str:
    if v is None:
        return "-"
    return f"Rp {v:,.0f}"


HEADER_FILL = "1a56db"   # blue header
ALT_FILL    = "f0f4ff"   # alternating row


# ─── Excel helpers ────────────────────────────────────────────────────────────

def _hdr_style(ws, row: int, cols: int, title: str) -> None:
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(bold=True, size=13, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    ws.row_dimensions[row].height = 22


def _col_hdr(ws, row: int, labels: list[str]) -> None:
    thin = Side(style="thin", color="AAAAAA")
    border = Border(bottom=thin)
    for c, lbl in enumerate(labels, 1):
        cell = ws.cell(row=row, column=c, value=lbl)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="374151")
        cell.border = border
        cell.alignment = Alignment(horizontal="center")


def _data_row(ws, row: int, values: list[Any], alt: bool = False) -> None:
    fill = PatternFill("solid", fgColor=ALT_FILL) if alt else None
    for c, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=c, value=val)
        if fill:
            cell.fill = fill
        if isinstance(val, (int, float, Decimal)):
            cell.alignment = Alignment(horizontal="right")
            cell.number_format = '#,##0'


# ─── Excel: Trial Balance ─────────────────────────────────────────────────────

def trial_balance_to_excel(data: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Trial Balance"

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 16

    _hdr_style(ws, 1, 5, f"NERACA SALDO  –  per {data['as_of']}")
    _col_hdr(ws, 2, ["Kode", "Nama Akun", "Debit", "Kredit", "Saldo"])

    for i, line in enumerate(data["lines"], 3):
        _data_row(ws, i, [
            line["code"],
            line["name"],
            float(line["total_debit"]),
            float(line["total_credit"]),
            float(line["balance"]),
        ], alt=(i % 2 == 0))

    total_row = len(data["lines"]) + 3
    ws.cell(total_row, 1, "TOTAL").font = Font(bold=True)
    ws.cell(total_row, 3, float(data["total_debit"])).font = Font(bold=True)
    ws.cell(total_row, 4, float(data["total_credit"])).font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── Excel: Profit & Loss ─────────────────────────────────────────────────────

def profit_loss_to_excel(data: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Laba Rugi"
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 18

    title = f"LAPORAN LABA RUGI  {data['date_from']} s/d {data['date_to']}"
    _hdr_style(ws, 1, 3, title)

    r = 2
    ws.cell(r, 1, "PENDAPATAN").font = Font(bold=True, size=11)
    r += 1
    _col_hdr(ws, r, ["Kode", "Akun", "Jumlah"])
    r += 1
    for i, line in enumerate(data["income"]):
        _data_row(ws, r, [line["code"], line["name"], float(line["amount"])], alt=(i % 2 == 0))
        r += 1
    ws.cell(r, 2, "Total Pendapatan").font = Font(bold=True)
    ws.cell(r, 3, float(data["total_income"])).font = Font(bold=True)
    r += 2

    ws.cell(r, 1, "BEBAN").font = Font(bold=True, size=11)
    r += 1
    _col_hdr(ws, r, ["Kode", "Akun", "Jumlah"])
    r += 1
    for i, line in enumerate(data["expense"]):
        _data_row(ws, r, [line["code"], line["name"], float(line["amount"])], alt=(i % 2 == 0))
        r += 1
    ws.cell(r, 2, "Total Beban").font = Font(bold=True)
    ws.cell(r, 3, float(data["total_expense"])).font = Font(bold=True)
    r += 2

    laba = ws.cell(r, 2, "LABA / RUGI BERSIH")
    laba.font = Font(bold=True, size=12)
    net = ws.cell(r, 3, float(data["net_profit"]))
    net.font = Font(bold=True, size=12, color="1a56db" if float(data["net_profit"]) >= 0 else "DC2626")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── Excel: Balance Sheet ─────────────────────────────────────────────────────

def balance_sheet_to_excel(data: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Neraca"
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 18

    _hdr_style(ws, 1, 3, f"NERACA  –  per {data['as_of']}")

    def _section(label: str, lines: list, total: Decimal, start_row: int) -> int:
        r = start_row
        ws.cell(r, 1, label).font = Font(bold=True, size=11)
        r += 1
        for i, line in enumerate(lines):
            _data_row(ws, r, [line["code"], line["name"], float(line["amount"])], alt=(i % 2 == 0))
            r += 1
        ws.cell(r, 2, f"Total {label}").font = Font(bold=True)
        ws.cell(r, 3, float(total)).font = Font(bold=True)
        return r + 2

    r = _section("ASET", data["assets"], data["total_assets"], 2)
    r = _section("KEWAJIBAN", data["liabilities"], data["total_liabilities"], r)
    r = _section("EKUITAS", data["equity"], data["total_equity"], r)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── PDF: base document ───────────────────────────────────────────────────────

def _base_doc(buf: io.BytesIO) -> SimpleDocTemplate:
    return SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=1.5 * cm, rightMargin=1.5 * cm,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm,
    )


_STYLES = getSampleStyleSheet()
_TITLE_STYLE = ParagraphStyle("kompak_title", parent=_STYLES["Title"], fontSize=14, spaceAfter=6)
_NORMAL = _STYLES["Normal"]

_TABLE_HEADER = TableStyle([
    ("BACKGROUND",  (0, 0), (-1, 0), colors.HexColor("#1a56db")),
    ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
    ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
    ("FONTSIZE",    (0, 0), (-1, -1), 8),
    ("ALIGN",       (0, 0), (-1, -1), "LEFT"),
    ("ALIGN",       (2, 1), (-1, -1), "RIGHT"),
    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f0f4ff")]),
    ("GRID",        (0, 0), (-1, -1), 0.3, colors.HexColor("#e5e7eb")),
    ("TOPPADDING",  (0, 0), (-1, -1), 3),
    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
])


# ─── PDF: Trial Balance ───────────────────────────────────────────────────────

def trial_balance_to_pdf(data: dict) -> bytes:
    buf = io.BytesIO()
    doc = _base_doc(buf)
    story = [
        Paragraph(f"NERACA SALDO – per {data['as_of']}", _TITLE_STYLE),
        Spacer(1, 0.3 * cm),
    ]

    rows = [["Kode", "Nama Akun", "Debit", "Kredit", "Saldo"]]
    for line in data["lines"]:
        rows.append([
            line["code"], line["name"],
            _fmt_decimal(Decimal(str(line["total_debit"]))),
            _fmt_decimal(Decimal(str(line["total_credit"]))),
            _fmt_decimal(Decimal(str(line["balance"]))),
        ])
    rows.append(["", "TOTAL",
                 _fmt_decimal(Decimal(str(data["total_debit"]))),
                 _fmt_decimal(Decimal(str(data["total_credit"]))), ""])

    t = Table(rows, colWidths=[2.5*cm, 8*cm, 3*cm, 3*cm, 3*cm])
    t.setStyle(_TABLE_HEADER)
    story.append(t)
    doc.build(story)
    return buf.getvalue()


# ─── PDF: Profit & Loss ───────────────────────────────────────────────────────

def profit_loss_to_pdf(data: dict) -> bytes:
    buf = io.BytesIO()
    doc = _base_doc(buf)
    story = [
        Paragraph(f"LAPORAN LABA RUGI  {data['date_from']} s/d {data['date_to']}", _TITLE_STYLE),
        Spacer(1, 0.3 * cm),
        Paragraph("<b>PENDAPATAN</b>", _NORMAL),
    ]

    income_rows = [["Kode", "Akun", "Jumlah"]]
    for line in data["income"]:
        income_rows.append([line["code"], line["name"], _rp(Decimal(str(line["amount"])))])
    income_rows.append(["", "Total Pendapatan", _rp(Decimal(str(data["total_income"])))])
    t = Table(income_rows, colWidths=[2.5*cm, 10*cm, 5*cm])
    t.setStyle(_TABLE_HEADER)
    story += [Spacer(1, 0.2*cm), t, Spacer(1, 0.4*cm)]

    story.append(Paragraph("<b>BEBAN</b>", _NORMAL))
    expense_rows = [["Kode", "Akun", "Jumlah"]]
    for line in data["expense"]:
        expense_rows.append([line["code"], line["name"], _rp(Decimal(str(line["amount"])))])
    expense_rows.append(["", "Total Beban", _rp(Decimal(str(data["total_expense"])))])
    t = Table(expense_rows, colWidths=[2.5*cm, 10*cm, 5*cm])
    t.setStyle(_TABLE_HEADER)
    story += [Spacer(1, 0.2*cm), t, Spacer(1, 0.5*cm)]

    net = Decimal(str(data["net_profit"]))
    color = "#1a56db" if net >= 0 else "#DC2626"
    story.append(Paragraph(
        f'<b><font color="{color}">LABA / RUGI BERSIH: {_rp(net)}</font></b>',
        ParagraphStyle("net", parent=_NORMAL, fontSize=11)
    ))

    doc.build(story)
    return buf.getvalue()


# ─── PDF: Balance Sheet ───────────────────────────────────────────────────────

def balance_sheet_to_pdf(data: dict) -> bytes:
    buf = io.BytesIO()
    doc = _base_doc(buf)
    story = [
        Paragraph(f"NERACA – per {data['as_of']}", _TITLE_STYLE),
        Spacer(1, 0.3 * cm),
    ]

    def _section_pdf(label: str, lines: list, total: str) -> None:
        story.append(Paragraph(f"<b>{label}</b>", _NORMAL))
        rows = [["Kode", "Akun", "Jumlah"]]
        for line in lines:
            rows.append([line["code"], line["name"], _rp(Decimal(str(line["amount"])))])
        rows.append(["", f"Total {label}", _rp(Decimal(str(total)))])
        t = Table(rows, colWidths=[2.5*cm, 10*cm, 5*cm])
        t.setStyle(_TABLE_HEADER)
        story.append(Spacer(1, 0.2*cm))
        story.append(t)
        story.append(Spacer(1, 0.4*cm))

    _section_pdf("ASET", data["assets"], data["total_assets"])
    _section_pdf("KEWAJIBAN", data["liabilities"], data["total_liabilities"])
    _section_pdf("EKUITAS", data["equity"], data["total_equity"])

    doc.build(story)
    return buf.getvalue()


# ─── Dispatch ─────────────────────────────────────────────────────────────────

EXCEL_EXPORTERS = {
    "trial-balance": trial_balance_to_excel,
    "profit-loss":   profit_loss_to_excel,
    "balance-sheet": balance_sheet_to_excel,
}

PDF_EXPORTERS = {
    "trial-balance": trial_balance_to_pdf,
    "profit-loss":   profit_loss_to_pdf,
    "balance-sheet": balance_sheet_to_pdf,
}

FILENAMES = {
    "trial-balance": "neraca_saldo",
    "profit-loss":   "laba_rugi",
    "balance-sheet": "neraca",
}


def export_report(report_type: str, data: dict, fmt: str) -> tuple[bytes, str, str]:
    """Return (content_bytes, filename, content_type)."""
    if fmt == "excel":
        fn = EXCEL_EXPORTERS.get(report_type)
        if not fn:
            raise ValueError(f"Excel export not supported for {report_type}")
        base = FILENAMES.get(report_type, report_type)
        return fn(data), f"{base}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    elif fmt == "pdf":
        fn = PDF_EXPORTERS.get(report_type)
        if not fn:
            raise ValueError(f"PDF export not supported for {report_type}")
        base = FILENAMES.get(report_type, report_type)
        return fn(data), f"{base}.pdf", "application/pdf"
    else:
        raise ValueError(f"Unsupported format: {fmt}")
